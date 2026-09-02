"""
Main Orchestration Script for Customer Scraping Automation.

Processes customer/seller IDs from multi-sheet Excel files (Merged Data 1, 2, 3) or text files,
executes API #1 and API #3 sequentially, evaluates D2C status, saves exclusively D2C ('Yes')
records to CSV with sequential Sr No, tracks all evaluated sellers in progress.json,
and automatically stops once 10,000 D2C records are collected (with full resumability).
"""

import argparse
from datetime import datetime
import json
import logging
from logging.handlers import RotatingFileHandler
import os
from pathlib import Path
import sys
import time
from typing import Any, Dict, Iterator, List, Optional, Set, Tuple, Union

# Ensure customer_scraper root is in sys.path
BASE_DIR = Path(__file__).resolve().parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from config.settings import (
    CHUNK_SIZE,
    DEFAULT_SCRAPE_LIMIT,
    DEFAULT_SELLER_ID,
    INPUT_FILE_PATH,
    INPUT_SHEET_NAMES,
    INPUT_SHEET_NAME,
    INPUT_COLUMN_NAMES,
    INPUT_COLUMN_NAME,
    LOG_FILE_PATH,
    OUTPUT_CSV_PATH,
    OUTPUT_DIR,
    OUTPUT_EXCEL_PATH,
    PROGRESS_FILE_PATH,
    SESSION_CONFIG_PATH,
)
from auth.auth_manager import AuthManager, AuthExpiredError
from api.api_client import APIClient, APIError
from scrapers.api1_scraper import API1Scraper
from scrapers.api2_scraper import API2Scraper
from scrapers.api3_scraper import API3Scraper
from scrapers.instagram_scraper import InstagramScraper
from excel.excel_writer import CSVWriter, ExcelWriter


def setup_logger() -> logging.Logger:
    """
    Configures unified logging to both file and console with formatted timestamps.
    """
    logger = logging.getLogger("customer_scraper")
    logger.setLevel(logging.INFO)

    # Avoid duplicate handlers if re-initialized
    if logger.handlers:
        return logger

    # Log format: [YYYY-MM-DD HH:MM:SS] [LEVEL] message
    formatter = logging.Formatter(
        "[%(asctime)s] [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )

    # File Handler (Rotating max 10MB, 5 backups)
    LOG_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)
    file_handler = RotatingFileHandler(
        LOG_FILE_PATH,
        maxBytes=10 * 1024 * 1024,
        backupCount=5,
        encoding="utf-8"
    )
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    # Console Handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    return logger


logger = setup_logger()


class ProgressTracker:
    """
    Tracks completed customer IDs in progress.json to enable deterministic resume.
    """

    def __init__(self, progress_path: Path = PROGRESS_FILE_PATH):
        self.progress_path = progress_path
        self.completed_ids: Set[str] = set()
        self.last_completed_id: str = ""
        self.last_sheet_name: str = ""
        self.last_row_index: int = 0
        self.load()

    def load(self) -> None:
        """Loads progress from disk if file exists."""
        if self.progress_path.exists():
            try:
                with open(self.progress_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.completed_ids = set(data.get("completed_customer_ids", []))
                    self.last_completed_id = data.get("last_completed_customer_id", "")
                    self.last_sheet_name = data.get("last_sheet_name", "")
                    self.last_row_index = data.get("last_row_index", 0)
                    logger.info("Loaded %d previously evaluated seller IDs from %s", len(self.completed_ids), self.progress_path.name)
            except Exception as e:
                logger.error("Error reading progress file (%s): %s", self.progress_path, str(e))

    def mark_completed(self, customer_id: str, sheet_name: str = "", row_index: int = 0) -> None:
        """
        Marks a customer as completed and immediately writes to progress.json.
        """
        clean_id = str(customer_id).strip()
        self.completed_ids.add(clean_id)
        self.last_completed_id = clean_id
        if sheet_name:
            self.last_sheet_name = sheet_name
        if row_index:
            self.last_row_index = row_index
        self._save()

    def sync_with_csv(self, csv_ids: Set[str], last_id: Optional[str] = None) -> None:
        """
        Synchronizes state with completed IDs and last completed ID found in the CSV dataset.
        """
        updated = False
        if csv_ids:
            before_count = len(self.completed_ids)
            self.completed_ids.update(csv_ids)
            if len(self.completed_ids) > before_count:
                logger.info("Synchronized progress: found %d completed IDs in output datasets.", len(self.completed_ids))
                updated = True

        if last_id and str(last_id).strip():
            clean_last = str(last_id).strip()
            if self.last_completed_id != clean_last:
                self.last_completed_id = clean_last
                updated = True

        if updated:
            self._save()

    # Backward compatibility alias
    sync_with_excel = sync_with_csv

    def is_completed(self, customer_id: str) -> bool:
        """Checks whether customer ID has already been evaluated."""
        return str(customer_id).strip() in self.completed_ids

    def _save(self) -> None:
        """Atomic write to progress JSON."""
        try:
            payload = {
                "completed_customer_ids": sorted(list(self.completed_ids)),
                "last_completed_customer_id": self.last_completed_id,
                "last_sheet_name": self.last_sheet_name,
                "last_row_index": self.last_row_index,
                "total_completed": len(self.completed_ids),
                "updated_at": datetime.now().isoformat(),
            }
            tmp_path = self.progress_path.with_suffix(".tmp")
            with open(tmp_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=2)
            tmp_path.replace(self.progress_path)
        except Exception as e:
            logger.error("Failed to save progress marker: %s", str(e))


def ensure_input_excel_exists(excel_path: Path, txt_path: Optional[Path] = None) -> None:
    """
    Ensures input Excel file exists with standard sheets and columns.
    If not, automatically creates template sheets: Merged Data 1, Merged Data 2, Merged Data 3.
    """
    if excel_path.exists():
        return

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        default_cols = ["Date", "Vertical", "Request ID", "Seller ID", "Total Listing"]

        # Try reading from text file if available
        ids_to_write = []
        if txt_path and txt_path.exists():
            with open(txt_path, "r", encoding="utf-8") as f:
                for line in f:
                    c_id = line.strip()
                    if c_id and not c_id.startswith("#"):
                        ids_to_write.append(c_id)

        if not ids_to_write:
            ids_to_write = ["d519f67b462d4e10", "218598a2b41c4bcd", "aaa30e788efa4f6c", "1111218ddd74492b"]

        for i, sheet_name in enumerate(INPUT_SHEET_NAMES):
            if i == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)

            ws.append(default_cols)
            if i == 0:
                for c_id in ids_to_write:
                    ws.append(["2026-08-28", "General", "REQ-100", str(c_id), "10"])

        excel_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(excel_path)
        logger.info(
            "Created input template at %s with %d sheets: %s",
            excel_path.name,
            len(INPUT_SHEET_NAMES),
            ", ".join(INPUT_SHEET_NAMES),
        )
    except Exception as e:
        logger.warning("Could not auto-create %s: %s", excel_path.name, str(e))


def is_valid_seller_id(val: Any) -> bool:
    """
    Validates whether a candidate string is a plausible Flipkart seller / customer ID.
    Rejects column headers, short strings (e.g. 'wsr'), keywords, and non-ID data.
    """
    if val is None:
        return False
    s = str(val).strip().lower()
    if len(s) < 8 or len(s) > 64:
        return False
    # Reject known non-seller keywords
    if s in (
        "seller_id", "seller id", "sellerid", "customer_id", "customer id",
        "customerid", "vertical", "wsr", "none", "null", "date", "request_id",
        "request id", "total_listing", "total listing", "account_name"
    ):
        return False
    return True


def stream_customer_ids(
    file_path: Path,
    sheet_names: Optional[List[str]] = None,
    column_names: Optional[List[str]] = None,
) -> Iterator[Tuple[str, int, str]]:
    """
    Streams (sheet_name, row_index, seller_id) sequentially across sheets
    using low-memory read_only streaming for large (10+ lakh row) workbooks.
    
    Yields:
        Tuple of (sheet_name: str, row_index: int, seller_id: str)
    """
    target_sheets = sheet_names or INPUT_SHEET_NAMES

    # Auto-create template if file is missing
    if not file_path.exists() and file_path.suffix.lower() in (".xlsx", ".xlsm", ".xltx"):
        txt_fallback = file_path.with_suffix(".txt")
        ensure_input_excel_exists(file_path, txt_path=txt_fallback)

    if not file_path.exists():
        logger.error("Input file not found at %s", file_path)
        return

    # 1. Handle Excel (.xlsx / .xlsm / .xltx) input with read_only streaming
    if file_path.suffix.lower() in (".xlsx", ".xlsm", ".xltx"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            # Determine list of sheets to read in order
            available_sheets = wb.sheetnames
            sheets_to_process = []

            for target in target_sheets:
                for s_name in available_sheets:
                    if s_name.strip().lower() == target.strip().lower():
                        sheets_to_process.append(s_name)
                        break

            # Fallback if configured sheet names aren't present
            if not sheets_to_process:
                sheets_to_process = available_sheets

            logger.info(
                "Streaming seller IDs from Excel '%s' across sheets: %s",
                file_path.name,
                sheets_to_process,
            )

            for s_name in sheets_to_process:
                ws = wb[s_name]
                seller_col_idx = None
                header_found = False

                for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    # Check top 5 rows for column header
                    if not header_found and row_idx <= 5:
                        # Pass 1: Prioritize exact 'Seller ID' / 'seller_id' / 'sellerid'
                        for col_idx, cell_val in enumerate(row):
                            if cell_val is not None:
                                col_clean = str(cell_val).strip().lower().replace(" ", "_").replace("-", "_")
                                if col_clean in ("seller_id", "sellerid", "seller_account_id"):
                                    seller_col_idx = col_idx
                                    header_found = True
                                    logger.info("Sheet '%s': Found primary 'Seller ID' header in column %d ('%s')", s_name, col_idx + 1, str(cell_val).strip())
                                    break

                        # Pass 2: Secondary check for 'customer_id' if seller_id not present
                        if not header_found:
                            for col_idx, cell_val in enumerate(row):
                                if cell_val is not None:
                                    col_clean = str(cell_val).strip().lower().replace(" ", "_").replace("-", "_")
                                    if col_clean in ("customer_id", "customerid"):
                                        seller_col_idx = col_idx
                                        header_found = True
                                        logger.info("Sheet '%s': Found secondary 'Customer ID' header in column %d ('%s')", s_name, col_idx + 1, str(cell_val).strip())
                                        break

                        if header_found:
                            continue

                    # If header wasn't found by row 6, auto-detect column containing valid 16-char IDs
                    if not header_found and row_idx > 5:
                        for col_idx, cell_val in enumerate(row):
                            if is_valid_seller_id(cell_val):
                                seller_col_idx = col_idx
                                header_found = True
                                logger.info("Sheet '%s': Auto-detected Seller ID in column %d", s_name, col_idx + 1)
                                break
                        if not header_found:
                            seller_col_idx = 3  # Default to 4th column ('Seller ID' in Date, Vertical, Request ID, Seller ID, Total Listing)
                            header_found = True

                    # Extract seller ID
                    if seller_col_idx is not None and len(row) > seller_col_idx:
                        val = row[seller_col_idx]
                        if val is not None:
                            if isinstance(val, float) and val.is_integer():
                                clean_id = str(int(val))
                            else:
                                clean_id = str(val).strip()

                            if is_valid_seller_id(clean_id):
                                yield (s_name, row_idx, clean_id)
                            elif clean_id and clean_id.lower() not in ("none", "null", ""):
                                logger.debug("Sheet '%s' Row %d: Skipped non-seller value '%s'", s_name, row_idx, clean_id)

            wb.close()
            return

        except Exception as e:
            logger.error("Failed to stream Excel input file %s: %s", file_path, str(e))
            txt_fallback = file_path.with_suffix(".txt")
            if txt_fallback.exists():
                logger.info("Falling back to reading from %s...", txt_fallback.name)
                yield from stream_customer_ids(txt_fallback)
            return

    # 2. Handle Plain Text (.txt) input
    row_idx = 0
    with open(file_path, "r", encoding="utf-8") as f:
        for line in f:
            row_idx += 1
            clean_line = line.strip()
            if clean_line and not clean_line.startswith("#") and is_valid_seller_id(clean_line):
                yield ("TextFile", row_idx, clean_line)


def read_customer_ids(file_path: Path) -> List[str]:
    """Helper to read all IDs into a list (for backwards compatibility / small files)."""
    return [c_id for _, _, c_id in stream_customer_ids(file_path)]


def main():
    parser = argparse.ArgumentParser(description="Flipkart Customer Scraping Automation")
    parser.add_argument("--refresh-session", action="store_true", help="Refresh Flipkart tab in Chrome and update session.json")
    parser.add_argument("--monitor-session", action="store_true", help="Run 10-minute browser tab keepalive refresh loop")
    parser.add_argument("--seller-id", type=str, default=DEFAULT_SELLER_ID, help=f"Specific seller ID for session refresh / keepalive (default: {DEFAULT_SELLER_ID})")
    parser.add_argument("--import-curl", type=str, help="Import session headers and cookies from a copied cURL command")
    parser.add_argument("--set-cookie", type=str, help="Set cookie string directly")
    parser.add_argument("--chunk-size", type=int, default=CHUNK_SIZE, help=f"Number of sellers per output batch file (default: {CHUNK_SIZE})")
    parser.add_argument("--limit", type=int, default=DEFAULT_SCRAPE_LIMIT, help=f"Target number of records to scrape and save (default: {DEFAULT_SCRAPE_LIMIT})")
    args = parser.parse_args()

    chunk_size = args.chunk_size or CHUNK_SIZE
    max_scrape_limit = args.limit or DEFAULT_SCRAPE_LIMIT

    # 1. Initialize Authentication Manager
    auth_manager = AuthManager(SESSION_CONFIG_PATH)

    # Handle standalone CLI commands
    if args.refresh_session:
        logger.info("Manual session refresh requested via --refresh-session.")
        success = auth_manager.refresh_session(seller_id=args.seller_id)
        if success:
            print(f"\n[+] Successfully refreshed and saved {len(auth_manager.cookies)} cookies to {SESSION_CONFIG_PATH.name}")
        else:
            print("\n[-] Session refresh could not be completed.")
        return

    if args.monitor_session:
        logger.info("Keep-alive monitoring requested via --monitor-session.")
        import asyncio
        asyncio.run(auth_manager.playwright_handler._async_keepalive_loop(seller_id=args.seller_id))
        return

    if args.import_curl:
        success = auth_manager.import_curl(args.import_curl)
        if success:
            print(f"[+] Successfully imported session into {SESSION_CONFIG_PATH.name}")
        else:
            print("[-] Could not parse cookies from provided cURL string.")
        return

    if args.set_cookie:
        auth_manager.set_cookie_string(args.set_cookie)
        auth_manager._save_to_file()
        print(f"[+] Cookie string saved to {SESSION_CONFIG_PATH.name}")
        return

    logger.info("==========================================")
    logger.info("START - Flipkart Customer Scraping Automation")
    logger.info("  Mode: API #1 (Details) + API #2 (Approvals & QnA) + API #3 (Contacts)")
    logger.info("  Save Policy: ALL processed records saved (with isD2C: Yes / No)")
    logger.info("  Target Limit: %d records", max_scrape_limit)
    logger.info("  Input File: %s", INPUT_FILE_PATH.name)
    logger.info("==========================================")

    # Verify and ensure valid session credentials on startup
    logger.info("Verifying active session status...")
    session_valid = auth_manager.ensure_valid_session(seller_id=args.seller_id)
    if not session_valid:
        logger.warning("Initial session check failed. Attempting automated browser extraction from Chrome...")
        try:
            auth_success = auth_manager.refresh_session(seller_id=args.seller_id)
            if not auth_success or not auth_manager.cookies:
                logger.error("Authentication required to proceed. Please ensure Chrome is open with remote debugging on port 9222 and logged in.")
                return
        except Exception as e:
            logger.error("Authentication error: %s", str(e))
            return

    api_client = APIClient(auth_manager)
    api1 = API1Scraper(api_client)
    api2 = API2Scraper(api_client)
    api3 = API3Scraper(api_client)
    insta_scraper = InstagramScraper()
    
    csv_writer = CSVWriter(output_dir=OUTPUT_DIR, chunk_size=chunk_size)
    excel_writer = csv_writer
    progress_tracker = ProgressTracker()

    # 2. Synchronize progress with existing CSV datasets
    completed_ids = csv_writer.get_completed_customer_ids()
    last_csv_id = csv_writer.get_last_completed_customer_id()
    progress_tracker.sync_with_csv(completed_ids, last_id=last_csv_id)

    # 3. Flush any pending unpersisted data if present
    csv_writer.flush_pending()

    # Calculate current sequential Sr No counter for saved records
    current_sr_no = csv_writer.get_current_customer_count() + 1

    total_saved_in_session = 0
    d2c_yes_count = 0
    d2c_no_count = 0
    total_evaluated_in_session = 0
    skipped_count = 0
    failed_count = 0

    logger.info("Already evaluated sellers in progress: %d", len(progress_tracker.completed_ids))
    logger.info("Current total rows in CSV: %d (Next Sr No: %d)", current_sr_no - 1, current_sr_no)
    logger.info("Target for this session: Collect %d records", max_scrape_limit)

    consecutive_auth_failures = 0

    try:
        for sheet_name, row_idx, customer_id in stream_customer_ids(INPUT_FILE_PATH):
            if max_scrape_limit and total_saved_in_session >= max_scrape_limit:
                logger.info(
                    "🎉 [TARGET ACHIEVED] Successfully processed and saved %d records in this session! All done!",
                    total_saved_in_session
                )
                break

            # Check if customer was already completed / evaluated
            if progress_tracker.is_completed(customer_id):
                skipped_count += 1
                if skipped_count % 1000 == 0:
                    logger.info("[RESUME SCAN] Skipped %d already-evaluated sellers (Current Sheet: %s, Row: %d)", skipped_count, sheet_name, row_idx)
                continue

            max_seller_retries = 3
            for seller_attempt in range(1, max_seller_retries + 1):
                try:
                    # Calculate current batch metrics and target paths
                    batch_num = ((current_sr_no - 1) // chunk_size) + 1
                    batch_pos = ((current_sr_no - 1) % chunk_size) + 1
                    target_csv = csv_writer.get_csv_path_for_sr(current_sr_no)

                    # Step 1: Execute API #1 (Customer & Seller Details)
                    api1_data = api1.get_seller_details(customer_id)
                    account_name = api1_data.get("account_name", "")
                    account_status = api1_data.get("account_status", "")
                    support_mgr = api1_data.get("support_manager", "No")
                    tier = api1_data.get("seller_tier", "")

                    # Step 2: Execute API #2 (Brand Approval, Actual Brand Count & QnA Questions)
                    api2_data = api2.get_brand_approval_details(customer_id)
                    approved_brand = api2_data.get("approved_brand", 0)
                    actual_brand_count = api2_data.get("actual_brand_count", 0)
                    request_id = api2_data.get("request_id", "")
                    brand_name = api2_data.get("brand_name", "")
                    brand_owner = api2_data.get("brand_owner", "")
                    document_type = api2_data.get("document_type", "")
                    brand_website_link = api2_data.get("brand_website_link", "")
                    brand_is_d2c = api2_data.get("brand_is_d2c", False)

                    # Step 3: Execute API #3 (Seller Contact Details & Unique Email)
                    api3_data = api3.get_seller_contacts(customer_id)
                    unique_email = api3_data.get("unique_email", "No")
                    is_email_d2c = str(unique_email).strip().lower() == "yes"

                    # Step 4: Search Instagram for Brand Name (with strict brand-in-URL validation)
                    instagram_url = ""
                    from scrapers.instagram_scraper import extract_instagram_url_from_string, is_brand_in_instagram_url

                    # 4a. Check if brand_website_link is directly an Instagram profile
                    if brand_website_link and "instagram.com" in str(brand_website_link).lower():
                        cand_url = extract_instagram_url_from_string(brand_website_link)
                        target_name = brand_name or account_name
                        if cand_url and target_name and is_brand_in_instagram_url(target_name, cand_url):
                            instagram_url = cand_url

                    # 4b. Search Instagram primarily using Brand Name
                    if not instagram_url and brand_name:
                        instagram_url = insta_scraper.search_instagram(brand_name) or ""

                    # 4c. Search using unique approved brands from API #2
                    if not instagram_url and api2_data.get("unique_brands"):
                        for brand_item in api2_data["unique_brands"]:
                            if brand_item and str(brand_item).strip().lower() not in (str(brand_name).strip().lower(), ""):
                                found_insta = insta_scraper.search_instagram(brand_item) or ""
                                if found_insta:
                                    instagram_url = found_insta
                                    break

                    # 4d. Fallback search using Account Name (with strict validation against brand_name/account_name)
                    if not instagram_url and account_name:
                        found_insta = insta_scraper.search_instagram(account_name) or ""
                        if found_insta:
                            target_name = brand_name or account_name
                            if is_brand_in_instagram_url(target_name, found_insta):
                                instagram_url = found_insta

                    # Final validation safeguard: If instagram_url is present, ensure target brand name is included
                    target_name = brand_name or account_name
                    if instagram_url and target_name and not is_brand_in_instagram_url(target_name, instagram_url):
                        logger.info("🚫 [Instagram Validation] URL '%s' does not contain brand '%s'. Skipping.", instagram_url, target_name)
                        instagram_url = ""

                    is_insta_d2c = bool(instagram_url and str(instagram_url).strip().lower() not in ("null", "none", "", "n/a", "na"))

                    # Multi-Criteria isD2C Evaluation:
                    # 1. Unique Email == 'Yes' OR
                    # 2. Document Type in ('BAL', 'TM') OR
                    # 3. Brand Website Link is available and valid OR
                    # 4. Instagram Profile is found
                    is_d2c_yes = bool(is_email_d2c or brand_is_d2c or is_insta_d2c)
                    is_d2c_str = "Yes" if is_d2c_yes else "No"

                    # Combine Results across all APIs and Scrapers
                    combined_record = {
                        **api1_data,
                        **api2_data,
                        **api3_data,
                        "instagram_url": instagram_url,
                        "unique_email": unique_email,
                        "unique_email_yes_no": unique_email,
                        "isD2C": is_d2c_str,
                        "is_d2c": is_d2c_str,
                    }

                    # Step 5: Save ALL processed records to CSV/Excel
                    save_success = csv_writer.append_customer(combined_record, sr_no=current_sr_no)
                    if save_success:
                        progress_tracker.mark_completed(customer_id, sheet_name=sheet_name, row_index=row_idx)
                        total_saved_in_session += 1
                        if is_d2c_yes:
                            d2c_yes_count += 1
                        else:
                            d2c_no_count += 1

                        logger.info(
                            "[Sheet: %s | Row: %d | Batch #%d (%d/%d)] ID: %s | Account: %s | Appr: %s | Act: %s | ReqID: %s | Brand: %s | BrOwner: %s | Doc: %s | Web: %s | Insta: %s | UniqEmail: %s | isD2C: %s -> SAVED TO CSV (Total Saved: %d/%d | D2C Yes: %d | Sr No: %d | File: %s)",
                            sheet_name, row_idx, batch_num, batch_pos, chunk_size, customer_id, account_name, approved_brand, actual_brand_count, request_id or "-", brand_name or "-", brand_owner or "-", document_type or "-", brand_website_link or "-", instagram_url or "-", unique_email, is_d2c_str, total_saved_in_session, max_scrape_limit, d2c_yes_count, current_sr_no, target_csv.name
                        )
                        if total_saved_in_session % 100 == 0:
                            logger.info(
                                "💾 [100-ROW CHECKPOINT] %d rows flushed to disk in %s (Current Sr No: %d, D2C Yes: %d, Non-D2C: %d).",
                                total_saved_in_session, target_csv.name, current_sr_no, d2c_yes_count, d2c_no_count
                            )
                        if current_sr_no % chunk_size == 0:
                            logger.info(
                                "🎉 [BATCH COMPLETED] Batch #%d (%d sellers) fully saved to %s!",
                                batch_num, chunk_size, target_csv.name
                            )
                        current_sr_no += 1
                    else:
                        logger.error("Failed to persist data for customer ID: %s", customer_id)

                    total_evaluated_in_session += 1
                    consecutive_auth_failures = 0
                    break

                except AuthExpiredError as auth_err:
                    consecutive_auth_failures += 1
                    logger.warning(
                        "⚠️ [AUTH EXPIRY] Session expired while processing %s (Attempt %d/%d). Clearing expired session and auto-refreshing...",
                        customer_id, seller_attempt, max_seller_retries
                    )
                    try:
                        auth_manager.clear_session()
                        refresh_ok = auth_manager.refresh_session(seller_id=DEFAULT_SELLER_ID)
                        if refresh_ok and auth_manager.cookies:
                            logger.info("✅ [AUTH RECOVERED] Successfully re-authenticated via Chrome! Continuing scraping...")
                            consecutive_auth_failures = 0
                            time.sleep(0.5)
                            continue  # Automatically retry this seller attempt immediately
                    except Exception as refresh_err:
                        logger.error("Automatic session refresh failed: %s", str(refresh_err))

                    # If multiple consecutive auth failures occur
                    if seller_attempt >= max_seller_retries:
                        if consecutive_auth_failures >= 3:
                            logger.warning(
                                "[AUTH RETRY LOOP] Multiple consecutive auth failures (%d). Backing off 5s for automated recovery...",
                                consecutive_auth_failures
                            )
                            time.sleep(5)
                            try:
                                auth_manager.clear_session()
                                if auth_manager.refresh_session(seller_id=DEFAULT_SELLER_ID):
                                    consecutive_auth_failures = 0
                                    logger.info("✅ [AUTH RECOVERED] Auto-refresh succeeded after backoff. Resuming pipeline.")
                            except Exception as re_err:
                                logger.error("Automated refresh retry error: %s", str(re_err))

                        failed_count += 1
                        logger.warning("Skipping customer %s after %d failed auth attempts. Continuing to next seller...", customer_id, max_seller_retries)
                        break

                except APIError as api_err:
                    failed_count += 1
                    logger.warning("[WARNING] API error for customer ID %s: %s. Continuing to next seller...", customer_id, str(api_err))
                    break

                except Exception as unexp_err:
                    failed_count += 1
                    logger.error("[ERROR] Unexpected error for customer ID %s: %s. Continuing to next seller...", customer_id, str(unexp_err))
                    break

    except KeyboardInterrupt:
        logger.warning("Scraper execution interrupted by user (KeyboardInterrupt). Shutting down cleanly...")
    finally:
        # Flush any pending records
        excel_writer.flush_pending()
        
        logger.info("==========================================")
        logger.info("Scraping Summary:")
        logger.info("  Total saved in this session:       %d / %d", total_saved_in_session, max_scrape_limit)
        logger.info("  - D2C ('Yes') records:             %d", d2c_yes_count)
        logger.info("  - Non-D2C ('No') records:          %d", d2c_no_count)
        logger.info("  Total evaluated in this session:   %d", total_evaluated_in_session)
        logger.info("  Skipped (already evaluated):       %d", skipped_count)
        logger.info("  Failed:                            %d", failed_count)
        logger.info("  Total evaluated sellers tracked:   %d", len(progress_tracker.completed_ids))
        logger.info("  Total rows across CSV datasets:    %d", current_sr_no - 1)
        logger.info("  Output Directory:                  %s", OUTPUT_DIR)
        logger.info("==========================================")


if __name__ == "__main__":
    main()
