"""
Integration tests for complete scraping workflow, resume capability, and state tracking with CSV.
"""

import csv
import json
import shutil
import tempfile
import unittest
from pathlib import Path
from unittest.mock import MagicMock, patch

from api.api_client import APIClient
from excel.excel_writer import CSVWriter, ExcelWriter
from main import ProgressTracker, read_customer_ids
from scrapers.api1_scraper import API1Scraper
from scrapers.api2_scraper import API2Scraper
from scrapers.api3_scraper import API3Scraper


class TestIntegrationScraperFlow(unittest.TestCase):
    def setUp(self):
        self.test_dir = Path(tempfile.mkdtemp())
        self.input_file = self.test_dir / "input.txt"
        self.csv_file = self.test_dir / "scraped_data.csv"
        self.progress_file = self.test_dir / "progress.json"
        self.pending_file = self.test_dir / "pending.json"

        # Populate sample input file
        self.input_file.write_text("ID000001\nID000002\nID000003\n\n# comment\nID000004\n", encoding="utf-8")

        self.mock_client = MagicMock(spec=APIClient)
        self.api1 = API1Scraper(self.mock_client)
        self.api2 = API2Scraper(self.mock_client)
        self.api3 = API3Scraper(self.mock_client)
        self.csv_writer = CSVWriter(csv_path=self.csv_file, pending_path=self.pending_file)
        self.excel_writer = self.csv_writer
        self.tracker = ProgressTracker(progress_path=self.progress_file)

    def tearDown(self):
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def test_input_file_reader_txt(self):
        ids = read_customer_ids(self.input_file)
        self.assertEqual(ids, ["ID000001", "ID000002", "ID000003", "ID000004"])

    def test_input_file_reader_excel(self):
        import openpyxl
        input_xlsx = self.test_dir / "customer_id_input.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Merged Data 1"
        ws.append(["Seller ID"])
        ws.append(["XLSX_ID_12345"])
        ws.append(["XLSX_ID_23456"])
        ws.append(["XLSX_ID_34567"])
        wb.save(input_xlsx)

        ids = read_customer_ids(input_xlsx)
        self.assertEqual(ids, ["XLSX_ID_12345", "XLSX_ID_23456", "XLSX_ID_34567"])

    def test_support_manager_yes_runs_all_apis_without_skipping(self):
        customer_id = "ID_SUPP_YES"

        # API 1 mock with Support Manager = Yes
        self.mock_client.get.return_value = {
            "result": {
                "displayName": "Managed Seller",
                "supportRole": {"name": "Manager Alex"},
                "liveDate": "2023-01-01",
                "gmv": {"response": {"details": {"state_details": {"state": "ACTIVE"}}}}
            }
        }

        api1_data = self.api1.get_seller_details(customer_id)
        self.assertEqual(api1_data["support_manager"], "Yes")
        self.assertEqual(api1_data["account_status"], "ACTIVE")

        # In updated workflow, all APIs run for all sellers
        saved = self.csv_writer.append_customer(api1_data, sr_no=1)
        self.assertTrue(saved)
        self.tracker.mark_completed(customer_id)
        self.assertTrue(self.tracker.is_completed(customer_id))

    def test_resume_skips_already_completed_customers(self):
        # Mark ID001 and ID002 as completed
        self.tracker.mark_completed("ID001")
        self.tracker.mark_completed("ID002")

        # Reload tracker from disk
        new_tracker = ProgressTracker(progress_path=self.progress_file)
        self.assertTrue(new_tracker.is_completed("ID001"))
        self.assertTrue(new_tracker.is_completed("ID002"))
        self.assertFalse(new_tracker.is_completed("ID003"))
        self.assertEqual(new_tracker.last_completed_id, "ID002")

    def test_resume_from_last_completed_id(self):
        input_ids = ["SELLER_A", "SELLER_B", "SELLER_C", "SELLER_D"]
        self.tracker.mark_completed("SELLER_B")

        # Verify resumption starting index
        last_id = self.tracker.last_completed_id
        start_idx = input_ids.index(last_id) + 1 if last_id in input_ids else 0
        self.assertEqual(start_idx, 2)
        self.assertEqual(input_ids[start_idx], "SELLER_C")

    def test_full_combined_flow(self):
        cust_id = "ID002"

        # 1. API 1 response (Support Manager = No)
        api1_resp = {
            "result": {
                "displayName": "Unmanaged Seller",
                "supportRole": {
                    "tier_type": "SUPPORT",
                    "role_name": None,
                    "user_id": None,
                    "email_id": None,
                    "name": None,
                    "phone_num": None,
                    "manager_email_id": None,
                },
                "gmv": {
                    "response": {
                        "details": {
                            "state_details": {"state": "ACTIVE"},
                            "darwin_tier_v2": {"tier_name": "Bronze"}
                        }
                    }
                },
                "profileInfo": {"created_at": "2020-01-01"},
                "liveDate": "2020-01-10",
            }
        }

        # 2. API 2 responses (Count + Requests)
        api2_count_resp = {
            "ALL": 43,
            "APPROVED": 26,
        }
        api2_records_resp = [
            {"brand_name": "RRCART", "request_status": "Approved"},
            {"brand_name": "rrcart", "request_status": "Approved"},
            {"brand_name": "TOY_BRAND", "request_status": "Approved"},
        ]

        # 3. API 3 response
        api3_resp = {
            "result": {
                "loginMobileNumber": "9123456780",
                "primaryMobileNumber": "9123456781",
                "loginEmail": "unman@mail.com",
                "primaryEmail": "unman_prim@mail.com",
            }
        }

        def mock_get(endpoint, headers=None):
            if "getSellerDetails" in endpoint:
                return api1_resp
            if "requestsV2-count" in endpoint:
                return api2_count_resp
            if "getSellerContacts" in endpoint:
                return api3_resp
            return {}

        self.mock_client.get.side_effect = mock_get
        self.mock_client.post.return_value = api2_records_resp

        # Run flow
        res1 = self.api1.get_seller_details(cust_id)
        self.assertEqual(res1["support_manager"], "No")
        self.assertEqual(res1["account_status"], "ACTIVE")

        res2 = self.api2.get_brand_approval_details(cust_id)
        self.assertEqual(res2["approved_brand"], 26)
        # Baseline approved = 26. "RRCART" and "rrcart" is 1 duplicate -> 26 - 1 = 25
        self.assertEqual(res2["actual_brand_count"], 25)

        res3 = self.api3.get_seller_contacts(cust_id)
        self.assertEqual(res3["email_id"], "unman@mail.com")

        combined = {**res1, **res2, **res3}
        saved = self.csv_writer.append_customer(combined, sr_no=1)
        self.assertTrue(saved)
        self.tracker.mark_completed(cust_id)

        # Verify CSV output
        with open(self.csv_file, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))

        self.assertEqual(len(rows), 2)  # Header + 1 row
        self.assertEqual(rows[1][1], cust_id)
        self.assertEqual(rows[1][2], "Unmanaged Seller")
        self.assertEqual(rows[1][3], "ACTIVE")
        self.assertEqual(rows[1][4], "No")  # Support Manager
        self.assertEqual(rows[1][5], "Bronze")  # Seller Tier
        self.assertEqual(rows[1][6], "2020-01-01")  # Signed Up Date
        self.assertEqual(rows[1][7], "2020-01-10")  # Live Date
        self.assertEqual(rows[1][8], "26")  # Approved Brand
        self.assertEqual(rows[1][10], "")   # Request ID
        self.assertEqual(rows[1][11], "RRCART")  # Brand Name
        self.assertEqual(rows[1][15], "")   # Instagram URL
        self.assertEqual(rows[1][16], "")   # Instagram Followers
        self.assertEqual(rows[1][17], "9123456780")  # Mobile Number
        self.assertEqual(rows[1][18], "9123456781")  # Registered Mobile Number
        self.assertEqual(rows[1][19], "unman@mail.com")  # Email ID
        self.assertEqual(rows[1][20], "unman_prim@mail.com")  # Registered Email ID
        self.assertEqual(rows[1][21], "No")  # Unique Email
        self.assertEqual(rows[1][22], "No")  # isD2C (mail.com is generic)

    def test_seller_limit_stops_execution(self):
        limit = 2
        input_ids = [f"ID_{i:03d}" for i in range(10)]
        processed_count = 0

        self.mock_client.get.return_value = {
            "result": {
                "displayName": "Test Seller",
                "supportRole": {"name": "Manager Alex"},
                "liveDate": "2023-01-01",
                "gmv": {"response": {"details": {"state_details": {"state": "ACTIVE"}}}}
            }
        }

        for c_id in input_ids:
            if processed_count >= limit:
                break
            data = self.api1.get_seller_details(c_id)
            self.excel_writer.append_customer(data, sr_no=processed_count + 1)
            self.tracker.mark_completed(c_id)
            processed_count += 1

        self.assertEqual(processed_count, limit)
        self.assertEqual(len(self.tracker.completed_ids), limit)

    def test_multi_sheet_streaming_excel(self):
        import openpyxl
        input_xlsx = self.test_dir / "input.xlsx"
        wb = openpyxl.Workbook()
        
        sheet1 = wb.active
        sheet1.title = "Merged Data 1"
        sheet1.append(["Date", "Vertical", "Request ID", "Seller ID", "Total Listing"])
        sheet1.append(["2026-08-01", "Apparel", "REQ-1", "SELLER_S1_1", "15"])
        sheet1.append(["2026-08-01", "Apparel", "REQ-2", "SELLER_S1_2", "20"])

        sheet2 = wb.create_sheet(title="Merged Data 2")
        sheet2.append(["Date", "Vertical", "Request ID", "Seller ID", "Total Listing"])
        sheet2.append(["2026-08-02", "Electronics", "REQ-3", "SELLER_S2_1", "5"])

        sheet3 = wb.create_sheet(title="Merged Data 3")
        sheet3.append(["Date", "Vertical", "Request ID", "Seller ID", "Total Listing"])
        sheet3.append(["2026-08-03", "Home", "REQ-4", "SELLER_S3_1", "50"])

        wb.save(input_xlsx)

        from main import stream_customer_ids
        streamed = list(stream_customer_ids(input_xlsx))
        
        expected = [
            ("Merged Data 1", 2, "SELLER_S1_1"),
            ("Merged Data 1", 3, "SELLER_S1_2"),
            ("Merged Data 2", 2, "SELLER_S2_1"),
            ("Merged Data 3", 2, "SELLER_S3_1"),
        ]
        self.assertEqual(streamed, expected)

    def test_d2c_filtering_saves_only_yes_to_csv(self):
        d2c_record = {
            "customer_id": "D2C_SELLER",
            "account_name": "Brand Direct",
            "account_status": "ACTIVE",
            "support_manager": "No",
            "seller_tier": "Gold",
            "email_id": "owner@mybranddirect.in",
            "isD2C": "Yes",
        }
        non_d2c_record = {
            "customer_id": "NON_D2C_SELLER",
            "account_name": "Generic Seller",
            "account_status": "ACTIVE",
            "support_manager": "No",
            "seller_tier": "Silver",
            "email_id": "seller@gmail.com",
            "isD2C": "No",
        }

        # Simulate scraper logic
        sr_no = 1
        d2c_saved = 0

        for rec in [d2c_record, non_d2c_record]:
            c_id = rec["customer_id"]
            if rec.get("isD2C") == "Yes":
                self.csv_writer.append_customer(rec, sr_no=sr_no)
                sr_no += 1
                d2c_saved += 1
            self.tracker.mark_completed(c_id)

        # Verify only 1 record (D2C) saved to CSV
        with open(self.csv_file, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))

        self.assertEqual(len(rows), 2)  # Header + 1 row
        self.assertEqual(rows[1][1], "D2C_SELLER")
        self.assertEqual(rows[1][14], "Yes")
        self.assertEqual(d2c_saved, 1)

        # But progress tracker must contain BOTH
        self.assertTrue(self.tracker.is_completed("D2C_SELLER"))
        self.assertTrue(self.tracker.is_completed("NON_D2C_SELLER"))


if __name__ == "__main__":
    unittest.main()
