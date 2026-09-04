"""
Persistence Layer for Customer Scraping Automation.

Handles writing scraped customer records to both CSV and Excel (.xlsx) formats,
with batch chunking (e.g. 1-500, 501-1000), locked-file retry recovery,
corrupted archive auto-repair, and durable pending buffer queues.
"""

import csv
import json
import logging
import os
from pathlib import Path
import shutil
import time
from typing import Any, Dict, List, Optional, Set, Tuple
import zipfile

from config.settings import (
    CHUNK_SIZE,
    CSV_COLUMNS,
    CSV_RETRY_INTERVAL,
    EXCEL_COLUMNS,
    GENERIC_EMAIL_DOMAINS,
    MAX_CSV_LOCK_RETRIES,
    OUTPUT_CSV_PATH,
    OUTPUT_DIR,
    OUTPUT_EXCEL_PATH,
    PENDING_FILE_PATH,
)

logger = logging.getLogger("customer_scraper")


class CSVWriter:
    """
    Manages writing scraped customer records to CSV and Excel workbooks
    with automatic batch chunking and file-lock recovery.
    """

    def __init__(
        self,
        csv_path: Optional[Path] = None,
        excel_path: Optional[Path] = None,
        pending_path: Optional[Path] = None,
        chunk_size: int = CHUNK_SIZE,
        output_dir: Optional[Path] = None,
    ):
        self.output_dir = Path(output_dir or OUTPUT_DIR)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.chunk_size = max(1, int(chunk_size or CHUNK_SIZE))
        self.csv_path = Path(csv_path) if csv_path else None
        self.excel_path = Path(excel_path) if excel_path else None
        self.pending_path = Path(pending_path or PENDING_FILE_PATH)

        # If explicit paths provided, ensure initial files exist
        if self.csv_path:
            self._ensure_csv_file_exists(self.csv_path)
        if self.excel_path:
            self._ensure_excel_file_exists(self.excel_path)

    def get_batch_range(self, sr_no: int) -> Tuple[int, int]:
        """
        Calculates the 1-indexed batch start and end Sr No for a given customer sequence number.
        Example with chunk_size=500:
            sr_no=1    -> (1, 500)
            sr_no=500  -> (1, 500)
            sr_no=501  -> (501, 1000)
        """
        sr = max(1, int(sr_no))
        batch_idx = (sr - 1) // self.chunk_size
        start_sr = (batch_idx * self.chunk_size) + 1
        end_sr = (batch_idx + 1) * self.chunk_size
        return start_sr, end_sr

    def get_excel_path_for_sr(self, sr_no: int) -> Path:
        """Returns the target Excel workbook path for the given Sr No."""
        if self.excel_path:
            return self.excel_path
        start_sr, end_sr = self.get_batch_range(sr_no)
        return self.output_dir / f"scraped_data_{start_sr}_to_{end_sr}.xlsx"

    def get_csv_path_for_sr(self, sr_no: int) -> Path:
        """Returns the target CSV path for the given Sr No."""
        if self.csv_path:
            return self.csv_path
        start_sr, end_sr = self.get_batch_range(sr_no)
        return self.output_dir / f"scraped_data_{start_sr}_to_{end_sr}.csv"

    def _ensure_csv_file_exists(self, file_path: Path) -> None:
        """Creates the target CSV file with formatted UTF-8-SIG headers if missing, or upgrades headers if column count changed."""
        file_path.parent.mkdir(parents=True, exist_ok=True)
        if not file_path.exists():
            try:
                with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    writer.writerow(CSV_COLUMNS)
                logger.debug("Initialized new CSV file at %s", file_path.name)
            except Exception as e:
                logger.error("Failed to initialize CSV file (%s): %s", file_path, str(e))
            return

        # Check if existing CSV has outdated columns (e.g. missing Brand Name or Instagram URL)
        try:
            with open(file_path, "r", newline="", encoding="utf-8-sig") as f:
                reader = list(csv.reader(f))

            if reader:
                existing_header = reader[0]
                if existing_header != CSV_COLUMNS:
                    logger.info("Migrating %s to %d-column schema...", file_path.name, len(CSV_COLUMNS))
                    header_map = {col.strip(): i for i, col in enumerate(existing_header)}
                    migrated_rows = [CSV_COLUMNS]
                    for row in reader[1:]:
                        new_row = []
                        for col_name in CSV_COLUMNS:
                            if col_name in header_map and header_map[col_name] < len(row):
                                new_row.append(row[header_map[col_name]])
                            else:
                                new_row.append("")
                        migrated_rows.append(new_row)

                    with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
                        writer = csv.writer(f)
                        writer.writerows(migrated_rows)
                    logger.info("Successfully upgraded %s to schema with %d columns.", file_path.name, len(CSV_COLUMNS))
        except Exception as mig_err:
            logger.debug("Schema migration check notice for %s: %s", file_path.name, str(mig_err))

    def _ensure_excel_file_exists(self, file_path: Path) -> None:
        """Creates the target Excel workbook with styled headers if missing."""
        if file_path.exists():
            return
        file_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Scraped Data"

            # Header row
            ws.append(CSV_COLUMNS)

            # Styling
            header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )

            for col_idx in range(1, len(CSV_COLUMNS) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = thin_border

            # Column widths (23 columns)
            col_widths = {
                1: 8,   # Sr No
                2: 20,  # Customer ID
                3: 25,  # Account Name
                4: 18,  # Account Status
                5: 16,  # Support Manager
                6: 14,  # Seller Tier
                7: 15,  # Signed Up Date
                8: 15,  # Live Date
                9: 16,  # Approved Brand
                10: 18, # Actual Brand Count
                11: 18, # Request ID
                12: 22, # Brand Name
                13: 16, # Brand Owner
                14: 16, # Document Type
                15: 32, # Brand Website Link
                16: 35, # Instagram URL
                17: 20, # Instagram Followers
                18: 18, # Mobile Number
                19: 24, # Registered Mobile Number
                20: 25, # Email ID
                21: 28, # Registered Email ID
                22: 14, # Unique Email
                23: 12, # isD2C
            }
            for col_idx, width in col_widths.items():
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = width

            ws.freeze_panes = "A2"
            wb.save(file_path)
            wb.close()
            logger.debug("Initialized new Excel workbook at %s", file_path.name)
        except ImportError:
            logger.debug("openpyxl not installed; skipping Excel formatting.")
        except Exception as e:
            logger.error("Failed to initialize Excel workbook (%s): %s", file_path, str(e))

    def _load_workbook_safe(self, file_path: Path) -> Optional[Any]:
        """Safely loads an Excel workbook, auto-recovering from corrupted/damaged archives."""
        try:
            import openpyxl
        except ImportError:
            return None

        if not file_path.exists():
            self._ensure_excel_file_exists(file_path)

        if file_path.exists():
            if file_path.stat().st_size == 0 or not zipfile.is_zipfile(file_path):
                logger.warning("Corrupted archive detected at %s. Re-creating workbook...", file_path.name)
                corrupt_bak = file_path.with_suffix(f".corrupt_{int(time.time())}.bak")
                try:
                    shutil.move(str(file_path), str(corrupt_bak))
                except Exception:
                    pass
                self._ensure_excel_file_exists(file_path)

        try:
            return openpyxl.load_workbook(file_path)
        except (KeyError, zipfile.BadZipFile, Exception) as e:
            logger.error("Could not load Excel workbook (%s): %s. Re-initializing...", file_path.name, str(e))
            corrupt_bak = file_path.with_suffix(f".corrupt_{int(time.time())}.bak")
            try:
                shutil.move(str(file_path), str(corrupt_bak))
            except Exception:
                pass
            self._ensure_excel_file_exists(file_path)
            try:
                return openpyxl.load_workbook(file_path)
            except Exception as e2:
                logger.critical("Critical failure loading workbook %s: %s", file_path.name, str(e2))
                return None

    def _append_rows_to_csv_with_retry(self, csv_file: Path, rows: List[List[Any]]) -> bool:
        """Appends rows to CSV with file-lock recovery."""
        if not rows:
            return True

        retries = 0
        while retries < MAX_CSV_LOCK_RETRIES:
            try:
                self._ensure_csv_file_exists(csv_file)
                with open(csv_file, "a", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    for row in rows:
                        writer.writerow(row)
                    f.flush()
                    try:
                        os.fsync(f.fileno())
                    except Exception:
                        pass
                logger.debug("Successfully appended %d row(s) to %s (flushed to disk)", len(rows), csv_file.name)
                return True
            except (PermissionError, OSError):
                retries += 1
                logger.warning(
                    "CSV file '%s' is locked (likely open in Excel). Save attempt %d/%d failed. Retrying in %ds...",
                    csv_file.name,
                    retries,
                    MAX_CSV_LOCK_RETRIES,
                    CSV_RETRY_INTERVAL,
                )
                time.sleep(CSV_RETRY_INTERVAL)

        logger.error("Failed to save CSV file '%s' after %d attempts. File lock not released.", csv_file.name, MAX_CSV_LOCK_RETRIES)
        return False

    def _append_rows_to_excel_with_retry(self, excel_file: Path, rows: List[List[Any]]) -> bool:
        """Appends rows to Excel workbook with file-lock recovery and styling."""
        if not rows:
            return True

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
        except ImportError:
            return True

        retries = 0
        while retries < MAX_CSV_LOCK_RETRIES:
            try:
                wb = self._load_workbook_safe(excel_file)
                if wb is None:
                    return False

                ws = wb.active
                body_font = Font(name="Calibri", size=10)
                body_align = Alignment(vertical="center")
                thin_border = Border(
                    left=Side(style="thin", color="E0E0E0"),
                    right=Side(style="thin", color="E0E0E0"),
                    top=Side(style="thin", color="E0E0E0"),
                    bottom=Side(style="thin", color="E0E0E0"),
                )

                for row_data in rows:
                    ws.append(row_data)
                    curr_row = ws.max_row
                    for c_idx in range(1, len(row_data) + 1):
                        cell = ws.cell(row=curr_row, column=c_idx)
                        cell.font = body_font
                        cell.alignment = body_align
                        cell.border = thin_border

                # Atomic save via temp file to avoid corruptions
                tmp_save_path = excel_file.with_suffix(f".tmp_{os.getpid()}_{int(time.time()*1000)}.xlsx")
                wb.save(tmp_save_path)
                wb.close()

                if excel_file.exists():
                    os.replace(tmp_save_path, excel_file)
                else:
                    tmp_save_path.replace(excel_file)

                logger.debug("Successfully appended %d row(s) to %s", len(rows), excel_file.name)
                return True

            except (PermissionError, OSError):
                retries += 1
                logger.warning(
                    "Excel file '%s' is locked. Save attempt %d/%d failed. Retrying in %ds...",
                    excel_file.name,
                    retries,
                    MAX_CSV_LOCK_RETRIES,
                    CSV_RETRY_INTERVAL,
                )
                time.sleep(CSV_RETRY_INTERVAL)
            except Exception as e:
                logger.error("Error writing to Excel workbook (%s): %s", excel_file.name, str(e))
                return False

        logger.error("Failed to save Excel file '%s' after %d attempts.", excel_file.name, MAX_CSV_LOCK_RETRIES)
        return False

    def load_pending_records(self) -> List[Dict[str, Any]]:
        """Loads unsaved records from the durable pending JSON file."""
        if not self.pending_path.exists():
            return []
        try:
            with open(self.pending_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, list):
                    return data
        except Exception as e:
            logger.error("Error reading pending results file (%s): %s", self.pending_path, str(e))
        return []

    def save_pending_records(self, pending_list: List[Dict[str, Any]]) -> None:
        """Persists pending customer records to disk atomically."""
        try:
            self.pending_path.parent.mkdir(parents=True, exist_ok=True)
            tmp_path = self.pending_path.with_suffix(f".tmp_{os.getpid()}")
            with open(tmp_path, "w", encoding="utf-8") as f:
                json.dump(pending_list, f, indent=2)
            os.replace(tmp_path, self.pending_path)
        except Exception as e:
            logger.error("Failed to write to pending file (%s): %s", self.pending_path, str(e))

    def clear_pending_records(self) -> None:
        """Clears pending records file upon successful persistence."""
        if self.pending_path.exists():
            try:
                self.pending_path.unlink()
            except Exception as e:
                logger.warning("Could not delete pending file: %s", str(e))

    def _clean_date_str(self, val: Any) -> str:
        """Strips time component from date string if present."""
        if val is None:
            return ""
        s = str(val).strip()
        if not s or s.lower() in ("null", "none"):
            return ""
        if "T" in s:
            return s.split("T")[0].strip()
        if " " in s:
            return s.split(" ")[0].strip()
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":
            return s[:10]
        return s

    def _format_customer_rows(self, data: Dict[str, Any], sr_no: Any) -> List[List[Any]]:
        """Formats customer scraped data into a tabular row matching the configured CSV_COLUMNS order."""
        customer_id = data.get("customer_id", "")
        account_name = data.get("account_name", "")
        account_status = data.get("account_status", "")
        support_manager = data.get("support_manager", "")
        seller_tier = data.get("seller_tier", "")
        signed_up_date = self._clean_date_str(data.get("signed_up_date", ""))
        live_date = self._clean_date_str(data.get("live_date", ""))
        approved_brand = data.get("approved_brand", "")
        actual_brand_count = data.get("actual_brand_count", "")
        request_id = data.get("request_id", "")
        brand_name = data.get("brand_name") or data.get("brand") or data.get("brandName") or ""
        brand_owner = data.get("brand_owner", "")
        document_type = data.get("document_type", "")
        brand_website_link = data.get("brand_website_link", "")
        instagram_url = data.get("instagram_url") or data.get("instagram") or ""
        instagram_followers = data.get("instagram_followers") or data.get("insta_followers") or ""
        mobile_number = data.get("mobile_number", "")
        registered_mobile = data.get("registered_mobile_number", "")
        email_id = data.get("email_id", "")
        registered_email = data.get("registered_email_id", "")
        unique_email = data.get("unique_email") or data.get("unique_email_yes_no", "")

        # Determine Unique Email ('Yes' or 'No')
        if not unique_email:
            unique_email = "No"
            for em in (email_id, registered_email):
                if em:
                    em_str = str(em).strip().lower()
                    if "@" in em_str:
                        dom = em_str.split("@")[-1].strip()
                        if dom and "." in dom and dom not in GENERIC_EMAIL_DOMAINS and dom not in ("null", "none"):
                            unique_email = "Yes"
                            break

        # Determine isD2C ('Yes' or 'No') based on 4 criteria:
        # 1. Unique Email == 'Yes'
        # 2. Document Type is BAL or TM
        # 3. Valid Brand Website Link
        # 4. Instagram Profile Found
        is_d2c = data.get("isD2C") or data.get("is_d2c")
        if not is_d2c:
            doc_type_clean = str(document_type).strip().upper()
            web_link_clean = str(brand_website_link).strip()
            is_valid_link = bool(
                web_link_clean
                and web_link_clean.lower() not in ("null", "none", "n/a", "na", "")
                and ("." in web_link_clean or "http" in web_link_clean.lower())
            )
            is_insta_found = bool(
                instagram_url
                and str(instagram_url).strip().lower() not in ("null", "none", "n/a", "na", "")
            )
            if unique_email == "Yes" or doc_type_clean in ("BAL", "TM") or is_valid_link or is_insta_found:
                is_d2c = "Yes"
            else:
                is_d2c = "No"

        return [[
            sr_no,
            customer_id,
            account_name,
            account_status,
            support_manager,
            seller_tier,
            signed_up_date,
            live_date,
            approved_brand,
            actual_brand_count,
            request_id,
            brand_name,
            brand_owner,
            document_type,
            brand_website_link,
            instagram_url,
            instagram_followers,
            mobile_number,
            registered_mobile,
            email_id,
            registered_email,
            unique_email,
            is_d2c,
        ]]

    def append_customer(self, customer_data: Dict[str, Any], sr_no: Any) -> bool:
        """
        Appends a single customer's scraped data to both CSV and Excel batch files.
        If locked or error, buffers to pending_results.json and retries.
        """
        # Buffer to pending queue first for durability
        pending = self.load_pending_records()
        pending.append({"sr_no": sr_no, "data": customer_data})
        self.save_pending_records(pending)

        # Flush all pending records
        success = self.flush_pending()
        if success:
            logger.info("Saved data successfully for customer ID: %s", customer_data.get("customer_id"))
        else:
            logger.warning("Pending save queued for record %s (%s)", str(sr_no), customer_data.get("customer_id"))

        return success

    def flush_pending(self) -> bool:
        """
        Writes all buffered pending records to both target CSV and Excel (.xlsx) files.
        """
        pending = self.load_pending_records()
        if not pending:
            return True

        # Group rows by target batch files based on sr_no
        csv_batches: Dict[Path, List[List[Any]]] = {}
        excel_batches: Dict[Path, List[List[Any]]] = {}

        for item in pending:
            item_sr = item.get("sr_no", 1)
            item_data = item.get("data", {})
            try:
                sr_int = int(item_sr) if str(item_sr).isdigit() else 1
            except Exception:
                sr_int = 1

            target_csv = self.get_csv_path_for_sr(sr_int)
            target_excel = self.get_excel_path_for_sr(sr_int)

            if target_csv not in csv_batches:
                csv_batches[target_csv] = []
            if target_excel not in excel_batches:
                excel_batches[target_excel] = []

            rows = self._format_customer_rows(item_data, item_sr)
            csv_batches[target_csv].extend(rows)
            excel_batches[target_excel].extend(rows)

        all_saved = True
        for target_csv, rows in csv_batches.items():
            csv_ok = self._append_rows_to_csv_with_retry(target_csv, rows)
            if not csv_ok:
                all_saved = False

        for target_excel, rows in excel_batches.items():
            excel_ok = self._append_rows_to_excel_with_retry(target_excel, rows)
            if not excel_ok:
                all_saved = False

        if all_saved:
            self.clear_pending_records()
            return True

        return False

    def get_completed_customer_ids(self) -> Set[str]:
        """
        Extracts all unique customer IDs persisted across all output CSV and Excel files.
        """
        completed_ids: Set[str] = set()

        # 1. Scan explicit CSV file if configured
        csv_files_to_check: Set[Path] = set()
        if self.csv_path and self.csv_path.exists():
            csv_files_to_check.add(self.csv_path)

        # 2. Scan all CSV files in output_dir
        if self.output_dir.exists():
            for f in self.output_dir.glob("*.csv"):
                csv_files_to_check.add(f)

        for csv_f in csv_files_to_check:
            try:
                with open(csv_f, "r", newline="", encoding="utf-8-sig") as f:
                    reader = csv.reader(f)
                    header = next(reader, None)
                    cust_col_idx = 1
                    if header:
                        for i, col in enumerate(header):
                            if col.strip().lower() in ("customer id", "customer_id", "seller id", "seller_id"):
                                cust_col_idx = i
                                break

                    for row in reader:
                        if row and len(row) > cust_col_idx:
                            c_id = str(row[cust_col_idx]).strip()
                            if c_id and c_id.lower() not in ("customer id", "customer_id", "none", "null", ""):
                                completed_ids.add(c_id)
            except Exception as e:
                logger.error("Could not read completed IDs from CSV (%s): %s", csv_f.name, str(e))

        # 3. Check Excel workbooks in output_dir if CSV had none
        if not completed_ids and self.output_dir.exists():
            try:
                import openpyxl
                for xlsx_f in self.output_dir.glob("*.xlsx"):
                    if xlsx_f.name.startswith("~") or ".tmp" in xlsx_f.name:
                        continue
                    try:
                        wb = openpyxl.load_workbook(xlsx_f, read_only=True, data_only=True)
                        ws = wb.active
                        rows_iter = ws.iter_rows(values_only=True)
                        header = next(rows_iter, None)
                        cust_col_idx = 1
                        if header:
                            for i, col in enumerate(header):
                                if col and str(col).strip().lower() in ("customer id", "customer_id", "seller id", "seller_id"):
                                    cust_col_idx = i
                                    break
                        for row in rows_iter:
                            if row and len(row) > cust_col_idx and row[cust_col_idx] is not None:
                                c_id = str(row[cust_col_idx]).strip()
                                if c_id and c_id.lower() not in ("customer id", "customer_id", "none", "null", ""):
                                    completed_ids.add(c_id)
                        wb.close()
                    except Exception:
                        pass
            except Exception:
                pass

        return completed_ids

    def get_current_customer_count(self) -> int:
        """Returns total count of unique completed customer IDs."""
        return len(self.get_completed_customer_ids())

    def get_last_completed_customer_id(self) -> str:
        """Returns the customer ID from the very last row written across all output CSV files."""
        csv_files = sorted(list(self.output_dir.glob("scraped_data_*.csv")), key=lambda p: p.stat().st_mtime)
        if not csv_files and self.csv_path and self.csv_path.exists():
            csv_files = [self.csv_path]

        last_id = ""
        for csv_f in reversed(csv_files):
            try:
                with open(csv_f, "r", newline="", encoding="utf-8-sig") as f:
                    reader = csv.reader(f)
                    header = next(reader, None)
                    cust_col_idx = 1
                    if header:
                        for i, col in enumerate(header):
                            if col.strip().lower() in ("customer id", "customer_id", "seller id", "seller_id"):
                                cust_col_idx = i
                                break
                    for row in reader:
                        if row and len(row) > cust_col_idx:
                            c_id = str(row[cust_col_idx]).strip()
                            if c_id and c_id.lower() not in ("customer id", "customer_id", "none", "null", ""):
                                last_id = c_id
                if last_id:
                    return last_id
            except Exception:
                pass
        return last_id


# Compatibility Aliases
ExcelWriter = CSVWriter
